"""
@ author István Hajdu at MTA TTK
https://github.com/hajduistvan/connectome_gan
"""
import yaml
from addict import Dict
from ast import literal_eval as make_tuple
import matplotlib.pyplot as plt
import numpy as np
import os
##############################################################################
#
# A simple example of some of the features of the XlsxWriter Python module.
#
# Copyright 2013-2018, John McNamara, jmcnamara@cpan.org
#
import openpyxl
import string


def get_excel_col_idx(i):
    # Indexing starts with 1=A, 2=B !!
    letters = string.ascii_uppercase
    if i <= len(letters):
        return letters[i - 1]
    else:
        raise NotImplementedError


excluded_keys = [
    'RELATIVE_DATA_FOLDER',
    'DATASET_PART',
    'SPLIT',
    'LOG_DIR',
    'SAVE_DIR',
    'SPLIT_LIMITS',
    'OUTLY_MIXER',
    'OUTLY_THRESHOLD',
]


def append_elements(dict, new_dict, excluded_keys, prefix):
    for k, v in dict.items():
        if not k in excluded_keys:
            if not type(v) == Dict:
                new_dict = {**new_dict, (prefix + '.' + k)[1:]: v}
            else:
                new_dict = append_elements(v, new_dict, excluded_keys, prefix + '.' + k)
    return new_dict


def save_params(CONFIG, result_dict):
    xls_filename = os.path.join(os.getcwd(), 'hyperparam_search.xlsx')
    param_dict = append_elements(CONFIG, {}, excluded_keys, '')

    sheet_name = param_dict['DATASET'] + '_' + param_dict['SUPERVISE_TYPE']
    write_exclude_keys = ['DATASET', 'SUPERVISE_TYPE']
    if os.path.isfile(xls_filename):
        wb = openpyxl.load_workbook(xls_filename)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
    if not sheet_name in wb.sheetnames:
        ws = wb.create_sheet()
        ws.title = sheet_name
    else:
        ws = wb[sheet_name]
    # Reading first column
    written_cols = tuple(ws.columns)
    written_keys = []

    if written_cols[0] == ():  # If key col is empty, fill it
        i = 1
        for key in sorted(result_dict.keys()):
            if key not in write_exclude_keys:
                ws['A' + str(i)] = key
                written_keys.append(key)
                i += 1
        for key in sorted(param_dict.keys()):
            if key not in write_exclude_keys:
                ws['A' + str(i)] = key
                written_keys.append(key)
                i += 1
    else:  # If not empty, fill with new elements if needed
        for cell in written_cols[0]:
            written_keys.append(cell.value)
        i = len(written_keys) + 1
        for key in sorted({**result_dict, **param_dict}.keys()):
            if not key in written_keys:
                ws['A' + str(i)] = key
                written_keys.append(key)
                i += 1
    num_col_to_write = len(written_cols) + 1
    col_letter = get_excel_col_idx(num_col_to_write)
    # Writing to new col
    for i, key in enumerate(written_keys):
        if not key in {**result_dict, **param_dict}.keys():
            val = ""
        else:
            val = {**result_dict, **param_dict}[key]
            if not type(val) == str and not val is None and val % 1:
                val = val - val % 0.001
        ws[col_letter + str(i + 1)] = val
    wb.save(xls_filename)
